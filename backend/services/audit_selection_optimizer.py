import math
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _stable_seed(text):
    return int(hashlib.md5(str(text).encode()).hexdigest(), 16) % (2**32)


def _safe_get_col(df_raw: pd.DataFrame, idx: int):
    if idx < df_raw.shape[1]:
        return df_raw.iloc[:, idx]
    return pd.Series([pd.NA] * len(df_raw))


def _select_units(audit, subset, target, md_value, risk_label, dept, log_lines):
    if subset.empty or md_value <= 0:
        log_lines.append(f"  {dept:<10} | {risk_label:<6} | No pool or MD<=0")
        return 0, 0

    num_units = max(math.floor(target / md_value), 0)
    if num_units == 0:
        log_lines.append(f"  {dept:<10} | {risk_label:<6} | Target too low for MD={md_value}")
        return 0, 0

    subset_sorted = subset.sort_values(by="Total Rating", ascending=False)

    if risk_label.lower() == "high" and num_units > 1:
        seed_value = _stable_seed(dept)
        top_n = math.ceil(num_units * 0.5)
        rem_n = num_units - top_n
        top_sel = subset_sorted.head(top_n)
        rem_pool = subset_sorted.iloc[top_n:]
        random_sel = rem_pool.sample(n=rem_n, random_state=seed_value) if rem_n > 0 and not rem_pool.empty else pd.DataFrame()
        final_sel = pd.concat([top_sel, random_sel]) if not random_sel.empty else top_sel
        log_lines.append(f"  {dept:<10} | {risk_label:<6} | {top_n} top + {rem_n} random x {md_value} days")
        log_lines.append(f"     🔹 Stable random seed for {dept}: {seed_value}")
    else:
        final_sel = subset_sorted.head(num_units)
        log_lines.append(f"  {dept:<10} | {risk_label:<6} | {num_units} units x {md_value} days")

    audit.loc[final_sel.index, "Selected"] = "Yes"
    audit.loc[final_sel.index, "Party days"] = md_value

    return len(final_sel), len(final_sel) * md_value


def run_audit_selection_optimizer(input_file: Path, output_dir: Path) -> Tuple[Path, Path, dict]:
    xls = pd.ExcelFile(input_file)
    params_raw = pd.read_excel(xls, sheet_name=0, header=None, dtype=object).dropna(how="all").reset_index(drop=True)
    audit_raw = pd.read_excel(xls, sheet_name=1, header=None, dtype=object).dropna(how="all").reset_index(drop=True)

    col1_numeric = pd.to_numeric(params_raw.iloc[:, 1], errors="coerce")
    if col1_numeric.dropna().empty:
        raise ValueError("No numeric values found in parameters column 1 to detect total mandays.")

    last_valid_idx = col1_numeric.dropna().index[-1]
    total_mandays = float(col1_numeric.iloc[last_valid_idx])
    params_data = params_raw.iloc[:last_valid_idx].reset_index(drop=True)

    if str(params_data.iloc[0, 0]).strip().lower().startswith("department"):
        params_data = params_data.iloc[1:].reset_index(drop=True)

    percent_col = pd.to_numeric(params_data.iloc[:, 1], errors="coerce")
    params_data = params_data.loc[percent_col.notna()].reset_index(drop=True)

    params = pd.DataFrame()
    params["Department"] = params_data.iloc[:, 0].astype(str).str.strip()
    params["Percentage"] = pd.to_numeric(params_data.iloc[:, 1], errors="coerce").fillna(0)
    params["HighDays"] = pd.to_numeric(params_data.iloc[:, 2], errors="coerce").fillna(0)
    params["MediumDays"] = pd.to_numeric(params_data.iloc[:, 3], errors="coerce").fillna(0)
    params["LowDays"] = pd.to_numeric(params_data.iloc[:, 4], errors="coerce").fillna(0)
    params["HighPct"] = pd.to_numeric(params_data.iloc[:, 5], errors="coerce").fillna(0)
    params["MedPct"] = pd.to_numeric(params_data.iloc[:, 6], errors="coerce").fillna(0)
    params["LowPct"] = pd.to_numeric(params_data.iloc[:, 7], errors="coerce").fillna(0)

    first_row_values = audit_raw.iloc[0].astype(str).str.lower().tolist()
    if any(keyword in first_row_values for keyword in ["department", "audit risk category", "s.no", "name of auditable audit"]):
        audit_full = pd.read_excel(xls, sheet_name=1, header=0, dtype=object)
    else:
        audit_full = audit_raw.copy()
        audit_full.columns = [f"Col_{i+1}" for i in range(audit_full.shape[1])]

    audit_full["Selected"] = "No"
    audit_full["Party days"] = 0

    audit = pd.DataFrame()
    audit["Department"] = _safe_get_col(audit_full, 3).astype(str).str.strip()
    audit["Section"] = _safe_get_col(audit_full, 5).astype(str).str.strip()
    audit["Audit Risk Category"] = _safe_get_col(audit_full, 8).astype(str).str.title().str.strip()
    audit["Total Rating"] = pd.to_numeric(_safe_get_col(audit_full, 7), errors="coerce").fillna(0)
    audit["Selected"] = audit_full["Selected"]
    audit["Party days"] = audit_full["Party days"]

    log_lines = []
    alloc_summary = []

    for _, row in params.iterrows():
        dept = str(row["Department"]).strip()
        pct = float(row["Percentage"])
        if pct == 0:
            continue

        dept_total = round(total_mandays * pct / 100)
        high_md, med_md, low_md = row["HighDays"], row["MediumDays"], row["LowDays"]
        high_pct, med_pct, low_pct = row["HighPct"], row["MedPct"], row["LowPct"]

        high_target = round(dept_total * high_pct / 100)
        med_target = round(dept_total * med_pct / 100)
        low_target = round(dept_total * low_pct / 100)

        dept_df = audit[audit["Department"].str.lower() == dept.lower()].copy()

        if dept_df.empty:
            log_lines.append(f"⚠️ {dept}: No matching audit units found.")
            alloc_summary.append((dept, dept_total, 0, 0))
            continue

        high_units, used_high = _select_units(audit, dept_df[dept_df["Audit Risk Category"] == "High"], high_target, high_md, "High", dept, log_lines)
        med_units, used_med = _select_units(audit, dept_df[dept_df["Audit Risk Category"] == "Medium"], med_target, med_md, "Medium", dept, log_lines)
        low_units, used_low = _select_units(audit, dept_df[dept_df["Audit Risk Category"] == "Low"], low_target, low_md, "Low", dept, log_lines)

        used_total = used_high + used_med + used_low
        utilization = round((used_total / dept_total) * 100, 1) if dept_total > 0 else 0
        alloc_summary.append((dept, dept_total, used_total, utilization))
        log_lines.append(f"{dept:10s} | Target={dept_total:5.0f} | Used={used_total:5.0f} | Util={utilization:5.1f}% | H:{high_units} M:{med_units} L:{low_units}")

    timestamp_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    base_name = input_file.stem
    output_file = output_dir / f"{base_name}_Results_{timestamp_str}.xlsx"
    log_file = output_dir / f"{base_name}_Log_{timestamp_str}.txt"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        audit_full.to_excel(writer, index=False, sheet_name="Selected Units")

        alloc_df = pd.DataFrame(alloc_summary, columns=["Department", "Mandays_Allocated", "Mandays_Used", "Utilization(%)"])
        alloc_df.to_excel(writer, index=False, sheet_name="Department Summary")

        section_analysis = (
            audit[audit["Selected"] == "Yes"]
            .groupby(["Section", "Audit Risk Category"])
            .size()
            .unstack(fill_value=0)
            .reindex(columns=["High", "Medium", "Low"], fill_value=0)
            .reset_index()
        )
        section_analysis.to_excel(writer, index=False, sheet_name="Section Analysis")

        dept_summary = (
            audit[audit["Selected"] == "Yes"]
            .groupby(["Department", "Audit Risk Category"])
            .size()
            .unstack(fill_value=0)
            .reindex(columns=["High", "Medium", "Low"], fill_value=0)
            .reset_index()
        )
        dept_summary["Total Selected Units"] = (
            dept_summary["High"] + dept_summary["Medium"] + dept_summary["Low"]
        )
        dept_summary.to_excel(writer, index=False, sheet_name="Dept Category Summary")

    wb = load_workbook(output_file)
    ws = wb["Selected Units"]
    sel_col = list(audit_full.columns).index("Selected") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=sel_col, max_col=sel_col):
        for cell in row:
            cell.fill = GREEN if cell.value == "Yes" else YELLOW
    wb.save(output_file)

    total_alloc = sum(x[1] for x in alloc_summary)
    total_used = sum(x[2] for x in alloc_summary)
    overall_util = round((total_used / total_alloc) * 100, 1) if total_alloc > 0 else 0
    selected_units = int((audit["Selected"] == "Yes").sum())
    risk_breakdown = (
        audit[audit["Selected"] == "Yes"]["Audit Risk Category"]
        .fillna("Unknown")
        .value_counts()
        .to_dict()
    )

    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "w", encoding="utf-8") as f:
        f.write(f"Audit Selection Optimizer Log\nTimestamp: {timestamp}\n\n")
        f.write(f"Input file used: {input_file}\n")
        f.write(f"Sheets found: {xls.sheet_names}\n")
        f.write(f"Total Mandays detected: {total_mandays}\n\n")
        for line in log_lines:
            f.write(line + "\n")
        f.write("\n" + "=" * 70 + "\n")
        f.write(f"🏁 TOTAL SUMMARY: Allocated={total_alloc} | Used={total_used} | Utilization={overall_util}%\n")

    summary_payload = {
        "total_mandays_allocated": total_alloc,
        "total_mandays_used": total_used,
        "overall_utilization": overall_util,
        "selected_units": selected_units,
        "risk_breakdown": risk_breakdown,
        "department_summary": alloc_df.to_dict(orient="records"),
        "section_analysis": section_analysis.to_dict(orient="records"),
    }

    return output_file, log_file, summary_payload

